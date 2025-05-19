#
#  excel_to_tsv  ファイル名
#

import argparse
import os
import glob
import pathlib
import openpyxl


def search_in_all_sheets(file_path: str):

    # Excelファイルを読み込む
    print(f"reading ... {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)

    results = []

    # すべてのシートを検索
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # 出力ファイル名生成
        file_name_without_ext = pathlib.Path(file_path).stem
        output_tsv_file_name = file_name_without_ext + "_" + sheet + ".tsv"

        # 出力バッファ
        outbuff = ""

        # 行を検索
        for row in ws.iter_rows():

            linebuff = ""

            # セルを検索
            for cell in row:
                if cell.value:
                    linebuff += str(cell.value) + "\t"
                else:
                    linebuff += "\t"
            # end of for cell

            outbuff += linebuff + "\n"

        # end of for row

        # ファイルへの書き出し
        print(f"writing ... {output_tsv_file_name}")
        with open(output_tsv_file_name, "wt", encoding='utf-8') as f:
            f.write(outbuff)

    # end of for sheet

    wb.close()

    return results


def list_files_with_extension(dirname: str, extension: str):
    # 指定されたフォルダ内の特定の拡張子を持つファイルのリストを返却
    files = glob.glob(os.path.join(dirname, f'**/*{extension}'), recursive=True)

    # 除外ファイル名チェック
    result_files = []
    for f in files:
        result_files.append(f)
    return result_files


def main():
    # 引数解析
    parser = argparse.ArgumentParser(description="excel to tsv ツール")

    # オプション引数
    parser.add_argument("first_arg", metavar='N', type=str, nargs='+', default="default", help="チェック対象のファイル")
    parser.add_argument("--utf8", action='store_true', help='stdoutをutf-8にする')
    parser.add_argument("--filter", help='対称ファイル名の中に含まれる文字列を指定')
    parser.add_argument("--charset", help="ファイルのcharset。デフォルトはutf-8。")

    # 引数解析
    args = parser.parse_args()

    extension = '.xlsx'

    # 引数を取得
    charset = 'utf-8'
    if args.charset:
        charset = args.charset

    if args.utf8:
        # windows環境のpythonにて、stdoutをutf8にする
        import io
        import sys
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding=charset)

    # ファイルリスト作成
    file_list = []
    if args.first_arg != "default":
        if isinstance(args.first_arg, list):
            for f in args.first_arg:
                file_list.append(f)
        else:
            file_list.append(args.first_arg)

    # 対象をチェックしていく
    for f in file_list:
        if os.path.isfile(f):
            search_in_all_sheets(f)
        if os.path.isdir(f):
            files = list_files_with_extension(f, extension)
            for f2 in files:
                if args.filter:
                    if args.filter not in f2:
                        continue
                search_in_all_sheets(f2)


if __name__ == "__main__":
    main()

#
# end of file
#
