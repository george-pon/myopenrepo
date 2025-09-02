#
#  excel_grep 検索文字列  ファイル名
#

import argparse
import os
import glob
import openpyxl


def search_in_all_sheets(file_path, search_text):
    # Excelファイルを読み込む
    print(f"reading ... {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)

    results = []

    # すべてのシートを検索
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and search_text in str(cell.value):
                    results.append((file_path, sheet, cell.row, cell.column, str(cell.value)))

    wb.close()

    return results


def list_files_with_extension(dirname: str, extension: str):
    # 指定されたフォルダ内の特定の拡張子を持つファイルのリストを返却
    files = glob.glob(os.path.join(dirname, f'**/*{extension}'), recursive=True)

    # 除外ファイル名チェック
    result_files = []
    for f in files:
        result_files.append(f)
    return result_files


def main():
    # 引数解析
    parser = argparse.ArgumentParser(description="htmlチェックツール")

    # オプション引数
    parser.add_argument("first_arg", metavar='N', type=str, nargs='+', default="default", help="チェック対象のファイル")
    parser.add_argument("--utf8", action='store_true', help='stdoutをutf-8にする')
    parser.add_argument("--filter", help='対称ファイル名の中に含まれる文字列を指定')
    parser.add_argument("--charset", help="ファイルのcharset。デフォルトはutf-8。")

    # 引数解析
    args = parser.parse_args()

    extension = '.xlsx'

    # 引数を取得
    charset = 'utf-8'
    if args.charset:
        charset = args.charset

    if args.utf8:
        # windows環境のpythonにて、stdoutをutf8にする
        import io
        import sys
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding=charset)

    # ファイルリスト作成
    file_list = []
    if args.first_arg != "default":
        if isinstance(args.first_arg, list):
            for f in args.first_arg:
                file_list.append(f)
        else:
            file_list.append(args.first_arg)

    # 第一引数は検索文字列
    search_text = file_list.pop(0)
    print(f"search_text is {search_text}")

    # 対象をチェックしていく
    result_list = []
    for f in file_list:
        if os.path.isfile(f):
            result_list += search_in_all_sheets(f, search_text)
        if os.path.isdir(f):
            files = list_files_with_extension(f, extension)
            for f2 in files:
                if args.filter:
                    if args.filter not in f2:
                        continue
                result_list += search_in_all_sheets(f2, search_text)

    # 結果の表示
    if result_list:
        print("検索結果:")
        for book, sheet, row, col, value in result_list:
            print(f"{book}  [{sheet}]  {row}:{col}  {value}")
    else:
        print("見つかりませんでした。")


if __name__ == "__main__":
    main()

#
# end of file
#
